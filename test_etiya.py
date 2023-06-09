#https://www.saucedemo.com/  --> test için 
#https://docs.pytest.org/en/7.1.x/example/simple.html  --> pytest dokumentasyon
#test prefixine sahip olmalı
from selenium import webdriver
from selenium.webdriver.common.by import By
import pytest
import openpyxl

# assert işlemi => testin sonucunu bir koşula bağlamak. selenium kodları yazılır assertten önce 
# debugging
# breakpoint -> kodu debug modda iken bekleteceğimiz satır. Kod akarken takılmak istenen kod satırını seçmeye yarar.

class TestEtiya():
    
    # her test öncesi bir method çalıştırmak isteyebilirsin bunu genellemek için setup method denir.
    def setup_method(self,method):  #setup method her test başlangıcı çalışacak fonksiyon.
        self.driver=webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")

    def readDataFromExcel(): # test olarak anlaşılmasın diye test prefixi ile başlatmadık.
        print('x')
    
    # teardown method her test sonunda çalışacak fonksiyon.
    def teardown_method(self):
        self.driver.quit()

    def readDataFromExcel(): # test olarak anlaşılmasın diye test prefixi ile başlatmadık.
        excelFile = openpyxl.load_workbook("data/testdeneme.xlsx")
        sheet = excelFile["Sheet1"] # sheet = excel de ki hangi sayfa ise onun adı.
        rows = sheet.max_row # sayfada ki maxsimum satır sayısını alır.
        data = []
        for i in range(2,rows):
            username = sheet.cell(i, 1).value # cell = hücre  # (satır=i alanında tutuluyor  , 1= sutün) 
            password = sheet.cell(i, 2).value
            data.append((username,password))
        return data

    # @pytest.mark.skip() # uygulamayı test ederken atlanması gereken bir alan varsa o kısmı teste dahil etmemek için kullanılmaktadır. sadece methoda değil classada verebiliriz.
    def test_header(self):
        logo = self.driver.find_element(By.CLASS_NAME,'login_logo')
        assert logo.text=="Swag Labs"
        # assert 1==1  # true 

    # test aynı anda birden fazla veri ile çalışsın istiyorsak.
    # python annotation  
    @pytest.mark.parametrize("username,password",readDataFromExcel())
    def test_login_invalid(self,username,password):
        loginBtn = self.driver.find_element(By.ID, 'login-button')
        usernameInput = self.driver.find_element(By.ID, "user-name")
        passwordInput = self.driver.find_element(By.ID,"password")
        usernameInput.click()
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn.click()
        errorContainer = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorContainer.text == "Epic sadface: Username and password do not match any user in this service"
